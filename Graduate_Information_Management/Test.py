from openpyxl import Workbook  # 导入操作 excel时所用的库
from openpyxl import load_workbook

IsJob = ['是', '否']


def Menu():  # 菜单主界面
    print('*' * 22)
    print("* 查看毕业生列表输入: 1 *")
    print("* 添加毕业生信息输入: 2 *")
    print("* 修改毕业生信息输入: 3 *")
    print("* 删除毕业生信息输入: 4 *")
    print("* 退出系统请输入     0 *")
    print('*' * 22)


def SelectStudent():
    print('*' * 25)
    print("*  查看所有信息列表输入:     1 *")
    print("*  按学号查询信息输入:       2 *")
    print("*  按年级查询信息输入:       3 *")
    print("*  按是否就业查询信息输入:    4 *")
    print("*  退出系统请输入            0 *")
    print('*' * 25)


def PrintStudentList():  # 打印excel文件中的数据
    def PrintAll():
        wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
        sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
        for row in sheet.rows:  # 循环每一行
            for cell in row:  # 循环每一行的单元格
                print(cell.value, end="      ")  # 打印出每一个单元格的数据
            print()
        print()

    def SelectById():
        id = int(input("请输入需要查询的学号："))
        if(CheckIdIsRight(id)==False):
            wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
            sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
            r = FindId(id)
            for i in range(1, 8):
                print(sheet.cell(1, i).value, end=" ")  # 打印出表头的信息
            print()
            for i in range(1, 8):
                print(sheet.cell(r, i).value, end=" ")  # 打印出该id对应的信息
            print()
        else:
            print("学号输入错误！")

    def SelectByGrade():
        wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
        sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
        grade = int(input("请输入要查询的年级："))
        for row in sheet.rows:  # 循环每一行
            for cell in row:  # 循环每一行的单元格
                if(cell.value==grade):  #找到年级符合的学生的信息
                    for cell in row:
                        print(cell.value, end=" ")  # 打印出这一行的信息
                    print()
        print()

    def SelectByIsJob():
        wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
        sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
        isjob = input("请输入要查询的学生是否已经就业 ：")
        if isjob in IsJob:  # 检查输入是否正确
            if isjob=='是':  # 如果要查询已经就业的学生
                for row in sheet.rows:  # 循环每一行
                    for cell in row:  # 循环每一行的单元格
                        if (cell.value == '是'):# 找到就业信息是 '是'的学生的那一行
                            for cell in row:
                                print(cell.value, end=" ")# 打印出这一行的信息
                            print()
                print()
            else:  # 查询 '否' 还没有就业的学生
                for row in sheet.rows:  # 循环每一行
                    for cell in row:  # 循环每一行的单元格
                        if (cell.value == '否'):# 找到就业信息是 '否'的学生的那一行的内容
                            for cell in row:
                                print(cell.value, end=" ")# 打印出这一行的信息
                            print()
                print()
        else:
            print("输入错误！")

    while 1:# 循环修改界面直到退出
        SelectStudent()
        a = int(input("请输入想要执行的操作："))

        while a:
            if a == 1:
                PrintAll()
                break
            if a == 2 :
                SelectById()
                break
            if a == 3:
                SelectByGrade()
                break
            if a == 4:
                SelectByIsJob()
                break
            if a < 0 or a > 4:
                print("输入错误！请重新输入：")
                break
        if a == 0:
            break


def CheckIdIsRight(id):  # 检查学号ID是否存在或格式不正确
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]:
        if id == column.value:
            print("学号存在")
            return False
    if id < 1000000000:
        print("学号格式不正确!")
        return False
    return True


def AddStudent():  # 添加学生信息模块
    r = []  # 建立一个新的列表 在将这个列表插入到excel表中
    ID = None
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    print("请输入学号：")
    id = int(input())
    if CheckIdIsRight(id) == False:  # 当学号已经存在时不能添加相同学号的学生信息
        while 1:
            print("请输入正确的学号！")
            id = int(input())
            if (CheckIdIsRight(id) == True):
                ID = id
                break
    if (CheckIdIsRight(id) == True):
        ID = id
    r.append(ID)
    name = input("请输入你的名字：")  # 添加姓名信息
    r.append(name)
    tell = input("请输入你的电话号码：")  # 添加电话号码信息
    while (1):
        if len(tell) != 11:
            print("电话号码格式不正确！请重新输入：")
            tell = input()
            if len(tell) == 11:
                print("输入成功！")
                break
        if len(tell) == 11:
            break
    r.append(tell)
    grade = int(input("请输入你的年级："))  # 添加年级信息
    while 1:
        if grade < 2000:  # 判断年级是否正确范围内
            print("年级输入不正确！请重新输入：")
            grade = int(input())
            if grade >= 2000:
                print("输入成功！")
                break
        if grade >= 2000:
            break
    r.append(grade)
    institute = input("请输入你的学院：")  # 添加学院信息
    r.append(institute)
    isjob = input("是否已经工作：输入 ：是或否！")  # 添加是否就业信息 当其 是 '是'时才能添加公司
    while 1:
        if (isjob in IsJob) == True:
            r.append(isjob)
            break
        else:
            print("输入错误请重新输入：")
            isjob = input()
    if r[5] == '是':  # 添加公司信息
        company = input("请输入你的公司名 ")
        r.append(company)
    else:
        company = '无'
        r.append(company)

    e_mail = input("请输入你的电子邮箱：")  # 添加邮箱信息
    r.append(e_mail)
    sheet.append(r)
    wb.close()
    wb.save('StudentList.xlsx')


def StudentPersonalMsg():  # 修改信息界面选择
    print('*' * 22)
    print("* 修改姓名请输入: 1 *")
    print("* 修改电话号码请输入: 2 *")
    print("* 修改是否就业请输入: 3 *")
    print("* 修改就业公司请输入: 4 *")
    print("* 退出修改请输入： 0 *")
    print('*' * 22)


def FindId(id):  # 在excel中找到该学号所在的行 用来修改该学生的信息
    i = 0
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]:
        i = i + 1
        if id == column.value:
            return i  # 返回行数


def ChangeStudent():  # 修改学生信息模块
    def changename(col, wb):  # 修改姓名
        name = input("请输入修改之后的名字：")
        sheet.cell(col, 2, name)
        wb.save('StudentList.xlsx')

    def changetell(col, wb):  # 修改电话号码 同样进行信息格式校对
        tell = input("请输入你的电话号码：")
        while 1:
            if len(tell) != 11:
                print("电话号码格式不正确！请重新输入：")
                tell = input()
                if len(tell) == 11:
                    print("输入成功！")
                    break
            if len(tell) == 11:
                break
        sheet.cell(col, 3, tell)
        wb.save('StudentList.xlsx')

    def changeisjob(col, wb):  # 修改是否就业状态信息
        IsJob = ['是', '否']
        isjob = input("是否已经工作：输入 ：是或否！")
        while (1):
            if ((isjob in IsJob) == True):
                sheet.cell(col, 6, isjob)
                wb.save('StudentList.xlsx')
                break
            else:
                print("输入错误请重新输入：")
                isjob = input()

    def changecompany(col, wb):  # 修改公司信息
        if (sheet.cell(col, 6).value == '是'):  # 判断是否就业
            company = input("请输入修改后的公司：")
            sheet.cell(col, 7, company)
            wb.save('StudentList.xlsx')
        else:
            print("请先修改是否就业：")
            changeisjob(col, wb)
            changecompany(col, wb)

    print("请输入你要修改的学生的学号：")
    id = int(input())
    if (CheckIdIsRight(id) == False):  # 检验学号是否存在
        print("学号正确！")
        col = FindId(id)
        wb = load_workbook('StudentList.xlsx')
        sheet = wb.active
        StudentPersonalMsg()
        while (1):
            a = int(input("请输入："))
            while a > 0:
                if a == 1:
                    changename(col, wb)
                    print("修改成功！")
                    break
                if a == 2:
                    changetell(col, wb)
                    print("修改成功！")
                    break
                if a == 3:
                    changeisjob(col, wb)
                    print("修改成功！")
                    break
                if a == 4:
                    changecompany(col, wb)
                    print("修改成功！")
                    break
                elif a > 4 or a < 0:
                    print("输入有误！")
                    break
            if a == 0:
                break
    else:
        print("学号输入错误！")


def DeleteStudent():  # 删除学生信息
    PrintStudentList()
    id = int(input("请输入要删除学生的学号！"))
    if (CheckIdIsRight(id) == False):  # 判断学号为id的学生是否在StudentList.xlsx中
        print("学号正确！")
        row = FindId(id)  # 查找其所在的行
        wb = load_workbook('StudentList.xlsx')
        sheet = wb.active
        sheet.delete_rows(row, 1)  # 删除该行
        wb.save('StudentList.xlsx')
    else:
        print("学号输入错误！")


def main():  # 主函数
    Menu()  # 先打印菜单
    while 1:
        a = int(input("请输入： "))
        while a:
            if a == 1:
                PrintStudentList()
                Menu()
                break
            if a == 2:
                AddStudent()
                Menu()
                break
            if a == 3:
                ChangeStudent()
                Menu()
                break
            if a == 4:
                DeleteStudent()
                Menu()
                break
            elif a > 4 or a < 0:
                print("输入有误！")
                break

        if a == 0:  # 按0退出进程
            print("系统已退出！")
            exit()


main()
