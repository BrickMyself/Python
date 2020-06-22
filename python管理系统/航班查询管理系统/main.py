from openpyxl import Workbook  # 导入操作 excel时所用的库
from openpyxl import load_workbook  # 用来打开已经存在的表


def Menu():  # 主菜单
    print(end=" "*55)
    print("-"*30)
    print(end=" " * 55)
    print("--- 欢迎使用航班查询系统-----")
    print(end=" " * 55)
    print("- 查询航班信息请输入:     1-")
    print(end=" " * 55)
    print("- 增加航班信息请输入:     2-")
    print(end=" " * 55)
    print("- 修改航班信息请输入:     3-")
    print(end=" " * 55)
    print("- 删除航班信息请输入:     4-")
    print(end=" " * 55)
    print("- 退出航班查询系统请输入: 0-")
    print(end=" " * 55)
    print("-"*30)


def InquiryMenu():  # 查询航班信息菜单
    print(end=" " * 55)
    print("-" * 30)
    print(end=" " * 55)
    print("- 查询所有航班请输入:      1-")
    print(end=" " * 55)
    print("- 按航班号查询请输入:      2-")
    print(end=" " * 55)
    print("- 按起飞地点查询请输入:    3-")
    print(end=" " * 55)
    print("- 按到达地点查询请输入:    4-")
    print(end=" " * 55)
    print("- 按航班班期查询请输入:    5-")
    print(end=" " * 55)
    print("- 按飞机型号查询请输入:    6-")
    print(end=" " * 55)
    print("- 退出航班查询系统请输入:  0-")
    print(end=" " * 55)
    print("-"*30)


def PrintFlightInfo():  # 打印出表的所有信息
    wb = load_workbook('FligntInfo.xlsx')  # 打开现在已经有的表
    sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
    for row in sheet.rows:  # 循环每一行
        for cell in row:  # 循环每一行的单元格
            print(cell.value, end="      ")  # 打印出每一个单元格的数据
        print()
    print()


def FindId(id):  # 在excel中找到该 id 所在的行 返回行数
    i = 0
    wb = load_workbook('FligntInfo.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]: # 循环航班号那一列的数据
        i = i + 1  # 计数器
        if id == column.value:   # 找到了返回
            return i  # 返回行数


def CheckIsRight(msg, x):  # 检查航班信息在表格的第x列是否存在
    wb = load_workbook('FligntInfo.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[x]:  # in 成员资格检查
        if msg == column.value:
            return False  # 信息存在返回一个 False
    return True  # 不存在则返回 True


def InquiryFlight():
    def ViewAllFlightInfo():  # 查看所有航班信息
        PrintFlightInfo()  # 调用打印所有信息功能

    def SelectByFlightId():  # 按航班号查询航班信息
        id = input("请输入你要查询的航班的航班号：")
        if CheckIsRight(id, 0) == False:  # 检验该航班号是存在
            wb = load_workbook('FligntInfo.xlsx')  # 打开现在已经有的表
            sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
            r = FindId(id)  # 找到航班号对应的一行
            for i in range(1, 10):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            for i in range(1, 10):
                print(sheet.cell(r, i).value, end="   ")  # 打印出该i行的信息
            print()
        else:
            print("航班号输入错误！")

    def SelectByFlightLoca(x,str1):  # 按起飞地点进行查询
        local = str1
        StartLocation = input("请输入你要查询的"+local+"地点：")
        if CheckIsRight(StartLocation, x) == False:  # 检查 起点或终点是否在表中
            wb = load_workbook('FligntInfo.xlsx')
            sheet = wb.active
            for i in range(1, 10):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            r = 0
            for column in list(sheet.columns)[x]:
                # 循环第 x 列
                r = r+1  # 计数器
                if (column.value == StartLocation):  # 找到起飞地点符合的航班的信息所在的行
                    for i in range(1, 10):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出r行对应的信息
                    print()
            print()
        else:
            print("没有"+StartLocation+"为"+local+"的航班")

    def SelectByFlightDate(x):
        date = float(input("请输入你要查询的班期："))
        if CheckIsRight(date, x) == False:  # 检查班期是否在表中
            wb = load_workbook('FligntInfo.xlsx')
            sheet = wb.active
            for i in range(1, 10):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            r = 0
            for column in list(sheet.columns)[x]:
                # 循环第 x 列
                r = r+1  # 计算第几行
                if (column.value == date):  # 找到班期符合的航班的信息所在行
                    for i in range(1, 10):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出对应的信息
                    print()
            print()
        else:
            print("没有"+str(date)+"的航班")

    def SelectByType(x):
        type = int(input("请输入你要查询的型号："))
        if CheckIsRight(type, x) == False:  # 检查型号是否在表中
            wb = load_workbook('FligntInfo.xlsx')
            sheet = wb.active
            for i in range(1, 10):
                print(sheet.cell(1, i).value, end="   ")
            print()
            r = 0
            for column in list(sheet.columns)[x]:
                # 循环第x列
                r = r+1  # 计算第几行
                if (column.value == type):  # 找到型号符合的航班的信息
                    for i in range(1, 10):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出型号为type对应的信息
                    print()

            print()
        else:
            print("没有"+str(type)+"的航班")

    while 1:
        InquiryMenu()
        a = int(input("请输入： "))
        while a:
            if a == 1:
                ViewAllFlightInfo()  # 查看所有信息
                break
            if a == 2:
                SelectByFlightId()  # 按航班号搜索
                break
            if a == 3:
                str1 = "起点"
                SelectByFlightLoca(1, str1)  # 按起飞地点查找
                break
            if a == 4:
                str1 = "终点"
                SelectByFlightLoca(2, str1)  # 按到达地点查找
                break
            if a == 5:
                SelectByFlightDate(3)  # 按班期查找
                break
            if a == 6:
                SelectByType(6)  # 按型号查找
                break
            elif a > 6 or a < 0:
                print("输入有误！")
                break
        if a == 0:
            break


def PushFlightInfo():
    r = []  # 建立一个新的列表 在将这个列表插入到excel表中
    ID = None
    wb = load_workbook('FligntInfo.xlsx')
    sheet = wb.active
    print("请输入航班号：")
    id = (input())
    if CheckIsRight(id, 0) == False:  # 当航班号已经存在时不能添加相同航班号的航班
        while 1:
            print("请输入正确的航班号！")
            id = input()
            if (CheckIsRight(id, 0) == True):
                ID = id
                break
    if (CheckIsRight(id, 0) == True):
        ID = id
    r.append(ID)  # 将输入的ID插入到列表中
    start = input("请输入该航班的起点：")  # 添加起点信息
    r.append(start)   # 将起点插入到列表中
    end = input("请输入该航班的终点：")  # 添加终点信息
    r.append(end)   # 将终点插入到列表中
    date = float(input("请输入班期："))  # 添加班期信息
    r.append(date)    # 将班期插入到列表中
    FlyTime = input("请输入该班期的起飞时间（格式为08:00:00):")  # 添加起飞时间信息
    r.append(FlyTime)   # 将起飞时间信息插入到列表中
    ArrivalTime = input("请输入该班期的到达时间（格式为08:00:00):")  # 添加到达时间信息
    r.append(ArrivalTime)   # 将到达时间信息插入到列表中
    type = int(input("请输入该航班的机型："))
    r.append(type)  # 将机型插入到列表中
    money1 = input("请输入该航班经济舱的价格,如500元：")
    r.append(money1)  # 将经济舱价格插入到列表中
    money2 = input("请输入该航班头等舱的价格,如500元: ")
    r.append(money2)  # 将头等舱价格插入到列表中
    sheet.append(r)    # 将整个列表插入到excel 表格中 即为插入一行数据
    wb.close()  # 关闭文件
    wb.save('FligntInfo.xlsx')  # 保存文件


def ModifyMenu():
    print(end=" " * 55)
    print("-"*30)
    print(end=" " * 55)
    print("- 修改航班号请输入:         1：-")
    print(end=" " * 55)
    print("- 修改起飞地点请输入:       2：-")
    print(end=" " * 55)
    print("- 修改到达地点请输入:       3：-")
    print(end=" " * 55)
    print("- 修改班期请输入:           4：-")
    print(end=" " * 55)
    print("- 修改起飞时间请输入:       5：-")
    print(end=" " * 55)
    print("- 修改到达时间请输入:       6：-")
    print(end=" " * 55)
    print("- 修改机型信息请输入:       7：-")
    print(end=" " * 55)
    print("- 修改经济舱机票价格请输入: 8：-")
    print(end=" " * 55)
    print("- 修改头等舱机票价格请输入: 9：-")
    print(end=" " * 55)
    print("- 退出修改航班信息请输入:   0：-")
    print(end=" " * 55)
    print("-"*30)

def PrintRowInfo(row, wb):  # 打印出表头和某一行的信息
    sheet = wb.active
    for i in range(1, 10):
        print(sheet.cell(1, i).value, end="   ")  # 打印出表头
    print()
    for i in range(1, 10):
        print(sheet.cell(row, i).value, end="   ")  # 打印出这一行的信息
    print()

def ModifyFlightInfo():
    def modifyId(row, wb):  # 修改航班号 # row 为其所在的信息的行 wb 是表格对象
        id = input("请输入修改之后的航班号：")
        sheet.cell(row, 1, id)
        wb.save('FligntInfo.xlsx')  # 保存
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifyStart(row, wb):  # 修改起点
        sloca = input("请输入修改之后的起点：")
        sheet.cell(row, 2, sloca)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifyEnd(row, wb):  # 修改终点
        eloca = input("请输入修改之后的终点：")
        sheet.cell(row, 3, eloca)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifyDate(row, wb):  # 修改班期
        date = float(input("请输入修改之后的班期,如6.19: "))
        sheet.cell(row, 4, date)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifyflytime(row, wb):  # 修改起飞时间
        flytime = input("请输入修改之后的起飞时间,如08:00:00: ")
        sheet.cell(row, 5, flytime)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifyarrivaltime(row, wb):  # 修改到达时间
        arritime = input("请输入修改之后的到达时间,如08:00:00: ")
        sheet.cell(row, 6, arritime)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifytype(row, wb):
        type = input("请输入修改之后的机型：")  # 修改机型
        sheet.cell(row, 7, type)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifymoney1(row, wb):  # 修改经济舱价格
        money1 = input("请输入修改之后的经济舱的价格(单位/元): ")
        sheet.cell(row, 8, money1)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def modifymoney2(row, wb):  # 修改头等舱价格
        money2 = input("请输入修改之后的头等舱的价格(单位/元): ")
        sheet.cell(row, 9, money2)
        wb.save('FligntInfo.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)
    print("请输入你要修改的航班的航班号：")
    id = (input())
    if (CheckIsRight(id, 0) == False):  # 检验航班号是否存在
        print("航班号正确！")
        row = FindId(id)  # 找到这个航班号对应哪一行的数据，对这一行的数据进行修改
        wb = load_workbook('FligntInfo.xlsx')  #  打开表格
        sheet = wb.active
        while 1:
            ModifyMenu()
            a = int(input("请输入："))
            while a > 0:
                if a == 1:
                    modifyId(row, wb)
                    break
                if a == 2:
                    modifyStart(row, wb)
                    break
                if a == 3:
                    modifyEnd(row, wb)
                    break
                if a == 4:
                    modifyDate(row, wb)
                    break
                if a == 5:
                    modifyflytime(row, wb)
                    break
                if a == 6:
                    modifyarrivaltime(row, wb)
                    break
                if a == 7:
                    modifytype(row, wb)
                    break
                if a == 8:
                    modifymoney1(row, wb)
                    break
                if a == 9:
                    modifymoney2(row, wb)
                    break
                elif a > 9 or a < 0:
                        print("输入有误！")
                        break
            if a == 0:
                break
    else:
        print("航班号输入错误！")


def DeleteFlightInfo():  # 删除航班信息
    PrintFlightInfo()
    id = input("请输入要删除的航班的航班号：")
    if (CheckIsRight(id, 0) == False):
        print("航班号正确！")
        row = FindId(id)  # 查找其所在的行
        wb = load_workbook('FligntInfo.xlsx')
        sheet = wb.active
        isdelete = input("是否删除该航班信息？输入是或否：")
        if isdelete == '是':
            sheet.delete_rows(row, 1)  # 删除该行
            wb.save('FligntInfo.xlsx')
            print("删除成功！")
        else:
            print("删除失败！")

    else:
        print("航班号输入错误！")


def MainFunction():  #主函数
    Menu()
    while 1:
        a = int(input("请输入： "))
        while a:
            if a == 1:
                InquiryFlight()  # 查询航班信息
                Menu()
                break
            if a == 2:
                PushFlightInfo()  # 添加航班信息
                Menu()
                break
            if a == 3:
                ModifyFlightInfo()  # 修改航班信息
                Menu()
                break
            if a == 4:
                DeleteFlightInfo()  # 删除航班信息
                Menu()
                break
            elif a > 4 or a < 0:
                print("输入有误！")
                break

        if a == 0:  # 按0退出进程
            print("系统已退出！")
            exit()


MainFunction()

