from openpyxl import Workbook  # openpy是python 中一种对excel表格操作的一个库
from openpyxl import load_workbook  # load_workbook 是用来操作已经存在的表


def Menu():  # 主菜单
    print("\033[36;1m")
    print("^"*150)
    print(end="^")
    print(end=" " * 50)
    print("^-^-----航班查询系统-----^-^")
    print(end="^")
    print(end=" " * 50)
    print("^ 增加航班信息请输入:     1^")
    print(end="^")
    print(end=" " * 50)
    print("^ 查看所有航班请输入:     2^")
    print(end="^")
    print(end=" " * 50)
    print("^ 查找航班信息请输入:     3^")
    print(end="^")
    print(end=" " * 50)
    print("^ 修改航班信息请输入:     4^")
    print(end="^")
    print(end=" " * 50)
    print("^ 删除航班信息请输入:     5^")
    print(end="^")
    print(end=" " * 50)
    print("^ 退出航班查询系统请输入: 0^")
    print(end="^")
    print("^"*150)


def FindId(id):  # 在excel中找到该 id 所在的行 返回行数
    i = 0
    wb = load_workbook('PlaneList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]: # 循环航班号那一列的数据
        i = i + 1  # 计数器
        if id == column.value:   # 找到了返回
            return i  # 返回行数


def CheckIsRight(msg, x):  # 检查航班信息在表格的第x列是否存在
    wb = load_workbook('PlaneList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[x]:  # in 成员资格检查
        if msg == column.value:
            return False  # 信息存在返回一个 False
    return True  # 不存在则返回 True


def AddPlaneMsg():
    r = []  # 建立一个新的列表 在将这个列表插入到excel表中,便于数据插入，安全且效率高
    ID = None
    wb = load_workbook('PlaneList.xlsx')  # 打开现在已经有的表
    sheet = wb.active  # 获取活跃也就是当前使用的第一个sheet
    print("请输入航班号：")
    id = input()
    if CheckIsRight(id, 0) == False:  # 当航班号已经存在时不能添加相同航班号的航班信息
        while 1:
            print("请输入正确的航班号！")
            id = input()
            if (CheckIsRight(id, 0) == True):
                ID = id
                break
    if (CheckIsRight(id, 0) == True):
        ID = id
    r.append(ID)  # 将输入的ID插入到列表中
    date = float(input("请输入航班的日期，如6月16输入：6.16："))  # 添加班航班的日期信息
    r.append(date)    # 将班期插入到列表中
    start = input("请输入该航班的起点地点：")  # 添加起点信息
    r.append(start)   # 将起点插入到列表中
    end = input("请输入该航班的目的地：")  # 添加终点信息
    r.append(end)  # 将终点插入到列表中
    flytime = input("请输入该班期的起飞时间（格式为08:00:00):")  # 添加起飞时间信息
    r.append(flytime)   # 将起飞时间信息插入到列表中
    arrivaltime = input("请输入该班期的到达时间（格式为08:00:00):")  # 添加到达时间信息
    r.append(arrivaltime)   # 将到达时间信息插入到列表中
    money = int(input("请输入该航班的票价："))
    r.append(money)  # 将票价插入到列表中
    type = int(input("请输入该航班的机型："))
    r.append(type)  # 将航班飞机的型号插入到列表中
    sheet.append(r)    # 将整个列表插入到excel 表格中 即为插入一行数据
    wb.close()  # 关闭文件
    wb.save('PlaneList.xlsx')  # 保存文件


def PrintPlaneList():  # 打印出航班信息列表的所有信息
    wb = load_workbook('PlaneList.xlsx')  # 打开现在已经有的表
    sheet = wb.active  # 获取当前活跃的表也就是第一个sheet
    for row in sheet.rows:  # 循环表每一行
        for cell in row:  # 循环每一行的单元格
            print(cell.value, end="      ")  # 打印出每一个单元格的数据
        print()
    print()


def FindPlaneMenu():
    print("\033[36;1m")
    print("^"*150)
    print(end="^")
    print(end=" " * 50)
    print("^-^---查找航班信息功能---^-^")
    print(end="^")
    print(end=" " * 50)
    print("^ 按航班号查询请输入:     1^")
    print(end="^")
    print(end=" " * 50)
    print("^ 按日期查询请输入:       2^")
    print(end="^")
    print(end=" " * 50)
    print("^ 按起飞地查询请输入:     3^")
    print(end="^")
    print(end=" " * 50)
    print("^ 按目的地查询请输入:     4^")
    print(end="^")
    print(end=" " * 50)
    print("^ 按票价信息查询请输入:   5^")
    print(end="^")
    print(end=" " * 50)
    print("^ 按飞机机型查询请输入:   6^")
    print(end="^")
    print(end=" " * 50)
    print("^ 退出查找返回主界面输入: 0^")
    print(end="^")
    print("^"*150)


def FindPlaneMsg():
    def FindByPlaneId():  # 按航班号查询航班信息
        id = input("请输入你要查询的航班的航班号：")
        if CheckIsRight(id, 0) == False:  # 检验该航班号是存在
            wb = load_workbook('PlaneList.xlsx')  # 打开现在已经有的表
            sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
            r = FindId(id)  # 找到航班号对应的一行
            for i in range(1, 9):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            for i in range(1, 9):
                print(sheet.cell(r, i).value, end="   ")  # 打印出该i行的信息
            print()
        else:
            print("航班号输入错误！")

    def FindByFlightDate():
        date = float(input("请输入你要查询的航班日期："))
        if CheckIsRight(date, 1) == False:  # 检查航班日期是否在表中
            wb = load_workbook('PlaneList.xlsx')
            sheet = wb.active
            for i in range(1, 9):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            r = 0
            for column in list(sheet.columns)[1]:
                # 循环第 x 列
                r = r+1  # 计算第几行
                if column.value == date:  # 找到班期符合的航班的信息所在行
                    for i in range(1, 9):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出对应的信息
                    print()
            print()
        else:
            print("没有"+str(date)+"的航班")

    def FindByPlaneLocation(x, str1):  # 按起点或终点地点进行查询
        local = str1
        location = input("请输入你要查询的"+local+"地点：")
        if CheckIsRight(location, x) == False:  # 检查 起点或终点是否在表中
            wb = load_workbook('PlaneList.xlsx')
            sheet = wb.active
            for i in range(1, 9):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            r = 0
            for column in list(sheet.columns)[x]:
                # 循环第 x 列
                r = r+1  # 计数器
                if (column.value == location):  # 找到起飞地点符合的航班的信息所在的行
                    for i in range(1, 9):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出r行对应的信息
                    print()
            print()
        else:
            print("没有"+location+"为"+local+"的航班")

    def FindByPlanePrice():
        price = int(input("请输入你要查找的机票价格："))
        if CheckIsRight(price, 6) == False:  # 检查机票价格是否在表中
            wb = load_workbook('PlaneList.xlsx')
            sheet = wb.active
            for i in range(1, 9):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            r = 0
            for column in list(sheet.columns)[6]:
                # 循环第 x 列
                r = r+1  # 计算第几行
                if column.value == price:  # 找到机票价格符合的航班的信息所在行
                    for i in range(1, 9):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出对应的信息
                    print()
            print()
        else:
            print("没有价格为"+str(price)+"的航班")

    def FindByType():
        type = int(input("请输入你要查找的航班的型号："))
        if CheckIsRight(type, 7) == False:  # 检查航班机型是否在表中
            wb = load_workbook('PlaneList.xlsx')
            sheet = wb.active
            for i in range(1, 9):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            r = 0
            for column in list(sheet.columns)[7]:
                # 循环第 x 列
                r = r+1  # 计算第几行
                if column.value == type:  # 找到机型符合的航班的信息所在行
                    for i in range(1, 9):
                        print(sheet.cell(r, i).value, end="   ")  # 打印出对应的信息
                    print()
            print()
        else:
            print("没有型号为"+str(type)+"的航班")

    while 1:
        FindPlaneMenu()
        a = int(input("请输入： "))
        while a:
            if a == 1:
                FindByPlaneId()  # 按航班号查询航班信息
                break
            if a == 2:
                FindByFlightDate()  # 按航班日期查找航班信息
                break
            if a == 3:
                str1 = "起点"
                FindByPlaneLocation(2, str1)  # 按起飞地点查找
                break
            if a == 4:
                str1 = "终点"
                FindByPlaneLocation(3, str1)  # 按到达地点查找
                break
            if a == 5:
                FindByPlanePrice()  # 按机票价格查找航班信息
                break
            if a == 6:
                FindByType()  # 按型号查找
                break
            elif a > 6 or a < 0:
                print("输入有误！")
                break
        if a == 0:
            break



def ModifyMenu():
    print("\033[36;1m")
    print("^"*150)
    print(end="^")
    print(end=" " * 50)
    print("-^-^----修改航班信息功能-----：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改航班号请输入:         1：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改班期请输入:           2：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改起飞地点请输入:       3：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改到达地点请输入:       4：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改起飞时间请输入:       5：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改到达时间请输入:       6：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改航班机票价格请输入:   7：-")
    print(end="^")
    print(end=" " * 50)
    print("- 修改航班机型信息请输入:   8：-")
    print(end="^")
    print(end=" " * 50)
    print("- 退出修改航班功能请输入:   0：-")
    print(end="^")
    print("-"*150)


def PrintRowInfo(row, wb):   # 打印出表头和某一行的信息
    sheet = wb.active
    for i in range(1, 9):
        print(sheet.cell(1, i).value, end="   ")  # 打印出表头
    print()
    for i in range(1, 9):
        print(sheet.cell(row, i).value, end="   ")  # 打印出这一行的信息
    print()


def ChangePlaneMsg():
    def ChangeId(row, wb):  # 修改航班号 # row 为其所在的信息的行 wb 是表格对象
        id = input("请输入修改之后的航班号：")
        sheet.cell(row, 1, id)
        wb.save('PlaneList.xlsx')  # 保存
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangeDate(row, wb):  # 修改班期
        date = float(input("请输入修改之后的班期,如6.19: "))
        sheet.cell(row, 2, date)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangeStart(row, wb):  # 修改起点
        start = input("请输入修改之后的起点：")
        sheet.cell(row, 3, start)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangeEnd(row, wb):  # 修改终点
        end = input("请输入修改之后的终点：")
        sheet.cell(row, 4, end)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangeBeginTime(row, wb):  # 修改起飞时间
        begin = input("请输入修改之后的起飞时间,如08:00:00: ")
        sheet.cell(row, 5, begin)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangeStopTime(row, wb):  # 修改到达时间
        stop = input("请输入修改之后的到达时间,如08:00:00: ")
        sheet.cell(row, 6, stop)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangePrice(row, wb):  # 修改机票价格
        money = int(input("请输入修改之后的经济舱的价格(单位/元): "))
        sheet.cell(row, 7, money)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    def ChangeType(row, wb):  # 修改机型
        type = int(input("请输入修改之后的机型："))  # 修改机型
        sheet.cell(row, 8, type)
        wb.save('PlaneList.xlsx')
        print("修改成功！")
        PrintRowInfo(row, wb)

    print("请输入你要修改的航班的航班号：")
    id = (input())
    if (CheckIsRight(id, 0) == False):  # 检验航班号是否存在
        print("航班号正确！")
        row = FindId(id)  # 找到这个航班号对应哪一行的数据，对这一行的数据进行修改
        wb = load_workbook('PlaneList.xlsx')  # 打开表格
        sheet = wb.active
        while 1:
            ModifyMenu()
            a = int(input("请输入："))
            while a > 0:
                if a == 1:
                    ChangeId(row, wb)
                    break
                if a == 2:
                    ChangeDate(row, wb)
                    break
                if a == 3:
                    ChangeStart(row, wb)
                    break
                if a == 4:
                    ChangeEnd(row, wb)
                    break
                if a == 5:
                    ChangeBeginTime(row, wb)
                    break
                if a == 6:
                    ChangeStopTime(row, wb)
                    break
                if a == 7:
                    ChangePrice(row, wb)
                    break
                if a == 8:
                    ChangeType(row, wb)
                    break
                elif a > 9 or a < 0:
                    print("输入有误！")
                    break
            if a == 0:
                break
    else:
        print("航班号输入错误！")


def DeletePlaneInfo():  # 删除航班信息
    PrintPlaneList()
    id = input("请输入要删除的航班的航班号：")
    if (CheckIsRight(id, 0) == False):
        print("航班号正确！")
        row = FindId(id)  # 查找其所在的行
        wb = load_workbook('PlaneList.xlsx')
        sheet = wb.active
        isdelete = input("是否删除该航班信息？输入是或否：")
        if isdelete == '是':
            sheet.delete_rows(row, 1)  # 删除该行
            wb.save('PlaneList.xlsx')
            print("删除成功！")
            PrintPlaneList()
        else:
            print("删除失败！")

    else:
        print("航班号输入错误！")


def MainFun():  #主函数
    Menu()
    while 1:
        a = int(input("请输入： "))
        while a:
            if a == 1:
                AddPlaneMsg()  # 添加航班信息
                Menu()
                break
            if a == 2:
                PrintPlaneList()  # 查看所有航班信息
                Menu()
                break
            if a == 3:
                FindPlaneMsg()  # 查找航班信息
                Menu()
                break
            if a == 4:
                ChangePlaneMsg()  # 修改航班信息
                Menu()
                break
            if a == 5:
                DeletePlaneInfo()  # 删除航班信息
                Menu()
                break

            elif a > 5 or a < 0:
                print("输入有误！")
                break

        if a == 0:  # 按0退出进程
            print("系统已退出！")
            exit()


MainFun()