from openpyxl import Workbook  # 导入操作 excel时所用的库
from openpyxl import load_workbook
IsJob = ['是', '否']


def Menu():  # 菜单主界面
    print(end=" " * 45)
    print('*' * 22)
    print(end=" "*45)
    print("* 查询毕业生信息输入: 1 *")
    print(end=" " * 45)
    print("* 添加毕业生信息输入: 2 *")
    print(end=" " * 45)
    print("* 修改毕业生信息输入: 3 *")
    print(end=" " * 45)
    print("* 删除毕业生信息输入: 4 *")
    print(end=" " * 45)
    print("* 查看排序统计请输入: 5 *")
    print(end=" " * 45)
    print("* 退出系统请输入：    0 *")
    print(end=" " * 45)
    print('*' * 22)


def SelectStudentMenu():
    print(end=" " * 45)
    print('*' * 25)
    print(end=" " * 45)
    print("*  查看所有信息请输入:   1 *")
    print(end=" " * 45)
    print("*  按学号查询信息输入:   2 *")
    print(end=" " * 45)
    print("*  按年级查询信息输入:   3 *")
    print(end=" " * 45)
    print("*  按是否就业查询输入:   4 *")
    print(end=" " * 45)
    print("*  退出查询功能请输入:   0 *")
    print(end=" " * 45)
    print('*' * 25)


def FindId(id):  # 在excel中找到该 id 所在的行 返回行数
    i = 0
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]: # 循环学号那一列的数据
        i = i + 1
        if id == column.value:   # 找到了返回
            return i  # 返回行数


def BubbleSort(l2):    # 冒泡排序对列表中的数据进行一个升序的排列
    for i in range(0, len(l2)):
        count = 0
        for j in range(1, len(l2)):
            if int(l2[j-1])>int(l2[j]):
                temp = l2[j]
                l2[j] = l2[j-1]
                l2[j-1] = temp
                count= count+1
        if(count == 0):    # 算法优点 当已经有序时就不再进行排序
            return l2
    return l2    # 返回排好序的 列表


def GetAllStudentByGadeOrMoney(x):
    l = []  # 建立一个空的列表  用于存放数据进行排序
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[x]:
        l.append(column.value) # 将薪资或年级插入到列表中 得到一个薪资或年级列表
    l2 = l[1:]  # 由于第一个是表头 将第二个位置以后的范围 拷贝给 l2
    l3 = BubbleSort(l2)  # 进行 薪资或年级排序     # 3 是排好序的列表
    i = 1
    l3.reverse()    # 进行一个反转列表 得到一个降序的列表
    while i < len(l3):     # 这是为了剔除列表中相同的元素
        if l3[i] == l3[i-1]:
            del l3[i-1]
        else:
            i = i+1
    for i in range(1, 10):
        print(sheet.cell(1, i).value, end="      ")  # 打印出表头的信息
    print()
    j = 0
    while j < len(l3):  # 按照排好序的列表对应的值 在excel中查找 打印出对应的信息
        for row in sheet.rows:  # 循环每一行
            for cell in row:  # 循环每一行的单元格
                if cell.value == l3[j]:  # 找到年级符合的学生的信息
                    for cell in row:
                        print(cell.value, end="       ")  # 打印出这一行的信息
                    print()
        j = j+1
    print()


def GetAllStudentById():  # 按学号排序打印出学生信息(升序)
    l = []  # 建立一个空的列表
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]:
        l.append(column.value)  # 将学号插入到列表中 得到一个学号列表
    l2 = l[1:]  # 由于第一个是表头 将第二个位置以后的范围 拷贝给 l2
    l3 = BubbleSort(l2)  # 进行 学号排序
    # 3 是排好序的列表
    for i in range(1, 10):
        print(sheet.cell(1, i).value, end="      ")  # 打印出表头的信息
    print()
    for i in range(0, len(l3)):  # 依次找到排好序的学号或年级对应的学生信息即可
        r = FindId(l3[i])  # 找到该行
        for j in range(1, 10):
            print(sheet.cell(r, j).value, end="       ")  # 打印出该id对应的信息
        print()


def PrintAll():
        wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
        sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
        for row in sheet.rows:  # 循环每一行
            for cell in row:  # 循环每一行的单元格
                print(cell.value, end="      ")  # 打印出每一个单元格的数据
            print()
        print()


def PrintStudentList():  # 打印excel文件中的数据

    def SelectById():
        id = int(input("请输入需要查询的学号："))
        if (CheckIdIsRight(id) == False):
            wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
            sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
            r = FindId(id)
            for i in range(1, 10):
                print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
            print()
            for i in range(1, 10):
                print(sheet.cell(r, i).value, end="   ")  # 打印出该id对应的信息
            print()
        else:
            print("学号输入错误！")

    def SelectByGrade():
        wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
        sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
        grade = int(input("请输入要查询的年级："))
        for i in range(1, 10):
            print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
        print()
        for row in sheet.rows:  # 循环每一行
            for cell in row:  # 循环每一行的单元格
                if(cell.value==grade):  #找到年级符合的学生的信息
                    for cell in row:
                        print(cell.value, end="   ")  # 打印出这一行的信息
                    print()
        print()

    def SelectByIsJob():
        wb = load_workbook('StudentList.xlsx')  # 打开现在已经有的表
        sheet = wb.active  # 获取当前活跃的表 也就是当前使用的表
        isjob = input("请输入要查询的学生是否已经就业 ：")
        if isjob in IsJob:  # 检查输入是否正确
            if isjob == '是':  # 如果要查询已经就业的学生
                for i in range(1, 10):
                    print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
                print()
                for row in sheet.rows:  # 循环每一行
                    for cell in row:  # 循环每一行的单元格
                        if (cell.value == '是'):  # 找到就业信息是 '是'的学生的那一行
                            for cell in row:
                                print(cell.value, end="   ")  # 打印出这一行的信息
                            print()
                print()
            else:  # 查询 '否' 还没有就业的学生
                for i in range(1, 10):
                    print(sheet.cell(1, i).value, end="   ")  # 打印出表头的信息
                print()
                for row in sheet.rows:  # 循环每一行
                    for cell in row:  # 循环每一行的单元格
                        if (cell.value == '否'):  # 找到就业信息是 '否'的学生的那一行的内容
                            for cell in row:
                                print(cell.value, end="   ")  # 打印出这一行的信息
                            print()
                print()
        else:
            print("输入错误！")

    while 1:  # 循环查找直到退出
        SelectStudentMenu()
        a = int(input("请输入想要执行的操作："))
        while a:
            if a == 1:
                PrintAll()
                break
            if a == 2 :
                SelectById()
                break
            if a == 3:
                SelectByGrade()
                break
            if a == 4:
                SelectByIsJob()
                break
            if a < 0 or a > 4:
                print("输入错误！请重新输入：")
                break
        if a == 0:
            break


def CheckIdIsRight(id):  # 检查学号ID是否存在或格式不正确
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    for column in list(sheet.columns)[0]:
        if id == column.value:
            print("学号存在")
            return False
    if id < 1000000000:
        print("学号格式不正确!")
        return False
    return True


def AddStudent():  # 添加学生信息模块
    r = []  # 建立一个新的列表 在将这个列表插入到excel表中
    ID = None
    wb = load_workbook('StudentList.xlsx')
    sheet = wb.active
    print("请输入学号：")
    id = int(input())
    if CheckIdIsRight(id) == False:  # 当学号已经存在时不能添加相同学号的学生信息
        while 1:
            print("请输入正确的学号！")
            id = int(input())
            if (CheckIdIsRight(id) == True):
                ID = id
                break
    if (CheckIdIsRight(id) == True):
        ID = id
    r.append(ID)  # 将输入的ID插入到列表中
    name = input("请输入你的名字：")  # 添加姓名信息
    r.append(name)   # 将姓名插入到列表中
    tell = input("请输入你的电话号码：")  # 添加电话号码信息
    while (1):
        if len(tell) != 11:
            print("电话号码格式不正确！请重新输入：")
            tell = input()
            if len(tell) == 11:
                print("输入成功！")
                break
        if len(tell) == 11:
            break
    r.append(tell)   # 将电话号码插入到列表中
    grade = int(input("请输入你的年级："))  # 添加年级信息
    while 1:
        if grade < 2000:  # 判断年级是否正确范围内
            print("年级输入不正确！请重新输入：")
            grade = int(input())
            if grade >= 2000:
                print("输入成功！")
                break
        if grade >= 2000:
            break
    r.append(grade)    # 将年级插入到列表中
    institute = input("请输入你的学院：")  # 添加学院信息
    r.append(institute)   # 将学院信息插入到列表中
    isjob = input("是否已经工作：输入 ：是或否！")  # 添加是否就业信息 当其 是 '是'时才能添加公司
    while 1:
        if (isjob in IsJob) == True:
            r.append(isjob)
            break
        else:
            print("输入错误请重新输入：")
            isjob = input()
    if r[5] == '是':  # 添加公司信息
        company = input("请输入你的公司名 ")
        r.append(company)
    else:
        company = '无'
        r.append(company)

    e_mail = input("请输入你的电子邮箱：")  # 添加邮箱信息
    r.append(e_mail)    # 将电子邮箱信息插入到列表中
    if r[5] == '是':
      money = input("请输入你的月薪：")  # 添加月薪信息
      r.append(money)     # 只有当已经就业时才可以添加月薪信息
    if r[5] == '否':
        money = 0     # 否则 月薪默认为 0
        r.append(money)
    sheet.append(r)    # 将整个列表插入到excel 表格中 即为插入一行数据
    wb.close()
    wb.save('StudentList.xlsx')


def StudentPersonalMsg():  # 修改信息界面选择
    print(end=" " * 45)
    print('*' * 23)
    print(end=" " * 45)
    print("* 修改学生姓名请输入: 1 *")
    print(end=" " * 45)
    print("* 修改电话号码请输入: 2 *")
    print(end=" " * 45)
    print("* 修改是否就业请输入: 3 *")
    print(end=" " * 45)
    print("* 修改就业公司请输入: 4 *")
    print(end=" " * 45)
    print("* 修改邮箱信息请输入: 5 *")
    print(end=" " * 45)
    print("* 修改月薪信息请输入: 6 *")
    print(end=" " * 45)
    print("* 退出修改请输入：    0 *")
    print(end=" " * 45)
    print('*' * 23)


def ChangeStudent():  # 修改学生信息模块
    def changename(row, wb):  # 修改姓名 # row 为其所在的信息的行 wb 是表格对象
        name = input("请输入修改之后的名字：")
        sheet.cell(row, 2, name)
        wb.save('StudentList.xlsx')

    def changetell(row, wb):  # 修改电话号码 同样进行信息格式校对
        tell = input("请输入你的电话号码：")
        while 1:
            if len(tell) != 11:
                print("电话号码格式不正确！请重新输入：")
                tell = input()
                if len(tell) == 11:
                    print("输入成功！")
                    break
            if len(tell) == 11:
                break
        sheet.cell(row, 3, tell)
        wb.save('StudentList.xlsx')

    def changeisjob(row, wb):  # 修改是否就业状态信息
        IsJob = ['是', '否']
        isjob = input("是否已经工作：输入 ：是或否！")
        while (1):
            if ((isjob in IsJob) == True):
                sheet.cell(row, 6, isjob)
                wb.save('StudentList.xlsx')
                break
            else:
                print("输入错误请重新输入：")
                isjob = input()

    def changecompany(row, wb):  # 修改公司信息
        if (sheet.cell(row, 6).value == '是'):  # 判断是否就业
            company = input("请输入修改后的公司：")
            sheet.cell(row, 7, company)
            wb.save('StudentList.xlsx')
        else:
            print("请先修改是否就业：")
            changeisjob(row, wb)
            changecompany(row, wb)

    def changemail(row, wb): # 修改学生邮箱信息
        mail = input("请输入修改之后的邮箱：")
        sheet.cell(row, 8, mail)
        wb.save('StudentList.xlsx')

    def changemoney(row , wb):  # 修改月薪信息
        if (sheet.cell(row, 6).value == '是'):  # 判断是否就业
          money = int(input("请输入修改之后的月薪："))
          sheet.cell(row, 9, money)
          wb.save('StudentList.xlsx')
        else:
            print("请先修改就业状态及就业公司！")
            changeisjob(row, wb)
            changecompany(row, wb)
            changemoney(row, wb)
    PrintAll()
    print("请输入你要修改的学生的学号：")
    id = int(input())
    if (CheckIdIsRight(id) == False):  # 检验学号是否存在
        print("学号正确！")
        row = FindId(id)
        wb = load_workbook('StudentList.xlsx')
        sheet = wb.active
        StudentPersonalMsg()
        while 1:
            a = int(input("请输入："))
            while a > 0:
                if a == 1:
                    changename(row, wb)
                    print("修改成功！")
                    break
                if a == 2:
                    changetell(row, wb)
                    print("修改成功！")
                    break
                if a == 3:
                    changeisjob(row, wb)
                    print("修改成功！")
                    break
                if a == 4:
                    changecompany(row, wb)
                    print("修改成功！")
                    break
                if a == 5:
                    changemail(row, wb)
                    print("修改成功！")
                    break
                if a == 6:
                    changemoney(row, wb)
                    print("修改成功！")
                    break
                elif a > 6 or a < 0:
                    print("输入有误！")
                    break
            if a == 0:
                break
    else:
        print("学号输入错误！")


def DeleteStudent():  # 删除学生信息
    PrintAll()
    id = int(input("请输入要删除学生的学号："))
    if (CheckIdIsRight(id) == False):  # 判断学号为id的学生是否在StudentList.xlsx中
        print("学号正确！")
        row = FindId(id)  # 查找其所在的行
        wb = load_workbook('StudentList.xlsx')
        sheet = wb.active
        isdelete = input("是否删除该学生信息？输入是或否：")
        if isdelete == '是':
            sheet.delete_rows(row, 1)  # 删除该行
            wb.save('StudentList.xlsx')
            print("删除成功！")
        else:
            print("删除失败！")

    else:
        print("学号输入错误！")


def SortMenu():
    print(end=" " * 45)
    print('*' * 30)
    print(end=" " * 45)
    print("*  按学号从小到大排序结果输入: 1 *")
    print(end=" " * 45)
    print("*  按年级从大到小排序结果输入: 2 *")
    print(end=" " * 45)
    print("*  按薪资从高到低排序结果输入: 3 *")
    print(end=" " * 45)
    print("*  退出此功能请输入:           0 *")
    print(end=" " * 45)
    print('*' * 30)


def SortData():
    SortMenu()
    while 1:
        a = int(input("请输入： "))
        while a:
            if a == 1:
                GetAllStudentById()
                SortMenu()
                break
            if a == 2:
                GetAllStudentByGadeOrMoney(3)
                SortMenu()
                break
            if a == 3:
                GetAllStudentByGadeOrMoney(8)
                SortMenu()
                break
            elif a > 2 or a < 0:
                print("输入有误！")
                break
        if a == 0:
            break


def main():  # 主函数
    Menu()  # 先打印菜单
    while 1:
        a = int(input("请输入： "))
        while a:
            if a == 1:
                PrintStudentList()
                Menu()
                break
            if a == 2:
                AddStudent()
                Menu()
                break
            if a == 3:
                ChangeStudent()
                Menu()
                break
            if a == 4:
                DeleteStudent()
                Menu()
                break
            if a == 5:
                SortData()
                Menu()
                break
            elif a > 5 or a < 0:
                print("输入有误！")
                break

        if a == 0:  # 按0退出进程
            print("系统已退出！")
            exit()


main()
